from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Sum, Avg, F, Q
from django.db.models.functions import Coalesce, ExtractMonth, ExtractQuarter, ExtractYear
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from books.models import (
    Book,
    BookRequest, 
    PurchasedBook, 
    BookSubscription, 
    LibrarySubscription
)
from payments.models import Order

User = get_user_model()


def calculate_median(values_list):
    """Calculates the median order value for a list of numbers."""
    if not values_list:
        return 0.0
    sorted_vals = sorted(values_list)
    n = len(sorted_vals)
    mid = n // 2
    if n % 2 == 0:
        return (sorted_vals[mid - 1] + sorted_vals[mid]) / 2.0
    return float(sorted_vals[mid])


def is_authorized_staff(user):
    if not user.is_authenticated:
        return False
    if user.is_staff or user.is_superuser:
        return True
    if hasattr(user, 'profile') and getattr(user.profile, 'role', None) in ['SUPERADMIN', 'LIBRARIAN', 'LIBMANAGER']:
        return True
    return False


def apply_date_filter(queryset, date_field, filter_type, start_date=None, end_date=None):
    now = timezone.now()

    if filter_type == 'custom' and start_date and end_date:
        try:
            s_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            e_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            return queryset.filter(**{f"{date_field}__date__gte": s_date, f"{date_field}__date__lte": e_date})
        except ValueError:
            pass

    if filter_type == 'today':
        return queryset.filter(**{f"{date_field}__date": now.date()})
    elif filter_type == 'this_week':
        return queryset.filter(**{f"{date_field}__gte": now - timezone.timedelta(days=7)})
    elif filter_type == 'this_month':
        return queryset.filter(**{f"{date_field}__month": now.month, f"{date_field}__year": now.year})
    elif filter_type == 'this_year':
        return queryset.filter(**{f"{date_field}__year": now.year})
    return queryset


def get_order_amount(order):
    """Helper function to reliably retrieve the non-zero amount paid for an order."""
    for attr in ['base_amount_inr', 'charged_amount', 'amount', 'total_amount', 'total', 'price']:
        val = getattr(order, attr, None)
        if val is not None and float(val) > 0:
            return float(val)
    return 0.0


def fetch_report_payload(report_type, date_filter='all', start_date=None, end_date=None, is_excel=False):
    standard_book_headers = ['Sr. No.', 'Book Number', 'Book Title', 'Author', 'Price', 'Status', 'Date']
    if is_excel:
        standard_book_headers += ['Publisher / Category', 'Shelf Location']

    # 0A. TOTAL USERS REPORT
    if report_type in ['total_users_report', 'total_users']:
        users_qs = User.objects.all().order_by('-date_joined')
        date_field = 'date_joined' if hasattr(User, 'date_joined') else None
        if date_field:
            users_qs = apply_date_filter(users_qs, date_field, date_filter, start_date, end_date)

        headers = ['Sr. No.', 'Full Name / Username', 'Email Address', 'Date Joined', 'Role / Status']
        if is_excel:
            headers += ['Last Login', 'Is Active', 'Is Superuser', 'Total Orders Placed']

        details = []
        for idx, usr in enumerate(users_qs, start=1):
            full_name = usr.get_full_name() or usr.username
            email = usr.email or 'N/A'
            raw_date = getattr(usr, 'date_joined', None)
            joined_str = raw_date.strftime('%b %d, %Y') if raw_date else 'N/A'

            role_str = 'STUDENT'
            if hasattr(usr, 'profile') and hasattr(usr.profile, 'role'):
                role_str = str(usr.profile.role).upper()
            elif usr.is_superuser:
                role_str = 'SUPERADMIN'
            elif usr.is_staff:
                role_str = 'STAFF'

            row = [idx, full_name, email, joined_str, role_str]

            if is_excel:
                last_login_raw = getattr(usr, 'last_login', None)
                last_login_str = last_login_raw.strftime('%b %d, %Y %H:%M') if last_login_raw else 'Never'
                is_active_str = 'Yes' if getattr(usr, 'is_active', True) else 'No'
                is_super_str = 'Yes' if getattr(usr, 'is_superuser', False) else 'No'
                orders_count = Order.objects.filter(user=usr, status__iexact='Success').count() if hasattr(Order, 'user') else 0
                
                row += [last_login_str, is_active_str, is_super_str, orders_count]

            details.append({'row': row})

        return {
            'title': 'Registered Users Directory',
            'summary': f'Total Registered Users: {len(details)}',
            'headers': headers,
            'details': details
        }

    # 0B. TOTAL BOOKS CATALOG REPORT
    elif report_type in ['total_books_report', 'total_books']:
        books_qs = Book.objects.all()
        date_field = next((df for df in ['created_at', 'added_at', 'publication_date'] if hasattr(Book, df)), None)
        if date_field:
            books_qs = apply_date_filter(books_qs, date_field, date_filter, start_date, end_date)

        headers = ['Sr. No.', 'Book ISBN / ID', 'Book Title', 'Author', 'Price', 'Status']
        if is_excel:
            headers += ['Category / Genre', 'Publisher', 'Publication Year', 'Views Count', 'Stock Quantity']

        details = []
        for idx, book in enumerate(books_qs, start=1):
            book_no = getattr(book, 'isbn', getattr(book, 'id', 'N/A'))
            title = getattr(book, 'title', 'Untitled')
            author = getattr(book, 'author', 'N/A')
            price_val = getattr(book, 'price', 0.0)
            price_str = f"₹{float(price_val or 0.0):,.2f}"
            status_str = 'AVAILABLE' if getattr(book, 'is_available', True) else 'OUT OF STOCK'

            row = [idx, str(book_no), title, str(author), price_str, status_str]

            if is_excel:
                category = str(getattr(book, 'category', getattr(book, 'genre', 'General')))
                publisher = str(getattr(book, 'publisher', 'N/A'))
                pub_year = str(getattr(book, 'publication_year', getattr(book, 'year', 'N/A')))
                views_count = getattr(book, 'views', getattr(book, 'views_count', 0))
                stock_qty = getattr(book, 'stock', getattr(book, 'quantity', 1 if getattr(book, 'is_available', True) else 0))

                row += [category, publisher, pub_year, views_count, stock_qty]

            details.append({'row': row})

        return {
            'title': 'Library Books Catalog',
            'summary': f'Total Catalog Titles: {len(details)}',
            'headers': headers,
            'details': details
        }

    # 1. SUBSCRIPTIONS REPORT
    elif report_type == 'subscriptions':
        headers = list(standard_book_headers)
        details = []
        sr_no = 1
        
        if hasattr(BookSubscription, 'objects'):
            b_qs = BookSubscription.objects.all()
            date_field = next((df for df in ['created_at', 'subscribed_at', 'date_joined'] if hasattr(BookSubscription, df)), None)
            if date_field:
                b_qs = apply_date_filter(b_qs, date_field, date_filter, start_date, end_date)

            for sub in b_qs:
                book_obj = getattr(sub, 'book', None)
                book_no = getattr(book_obj, 'isbn', getattr(book_obj, 'id', 'N/A')) if book_obj else 'N/A'
                book_title = getattr(book_obj, 'title', 'Book Access') if book_obj else 'Book Subscription'
                author = getattr(book_obj, 'author', 'N/A') if book_obj else 'N/A'
                price = f"₹{float(getattr(book_obj, 'price', 0.0) or 0.0):,.2f}" if book_obj and hasattr(book_obj, 'price') else 'N/A'
                raw_date = getattr(sub, date_field, None) if date_field else None
                date_str = raw_date.strftime('%b %d, %Y') if raw_date else 'N/A'

                row = [sr_no, str(book_no), book_title, str(author), price, 'ACTIVE', date_str]

                if is_excel:
                    category = str(getattr(book_obj, 'category', 'General')) if book_obj else 'N/A'
                    shelf = str(getattr(book_obj, 'shelf_location', 'Digital Shelf')) if book_obj else 'N/A'
                    row += [category, shelf]

                details.append({'row': row})
                sr_no += 1

        if hasattr(LibrarySubscription, 'objects'):
            l_qs = LibrarySubscription.objects.all()
            date_field = next((df for df in ['created_at', 'subscribed_at', 'date_joined'] if hasattr(LibrarySubscription, df)), None)
            if date_field:
                l_qs = apply_date_filter(l_qs, date_field, date_filter, start_date, end_date)

            for sub in l_qs:
                raw_date = getattr(sub, date_field, None) if date_field else None
                date_str = raw_date.strftime('%b %d, %Y') if raw_date else 'N/A'

                row = [sr_no, 'LIB-PASS', 'Library Membership', 'N/A', 'N/A', 'ACTIVE', date_str]

                if is_excel:
                    row += ['Library Wide Access', 'Main Reading Hall']

                details.append({'row': row})
                sr_no += 1

        return {
            'title': 'Subscription Analytics',
            'summary': f'Total Subscriptions: {len(details)}',
            'headers': headers,
            'details': details
        }

    # 2. ORDERS REPORT (WITH MEAN & MEDIAN)
    elif report_type == 'orders':
        orders_filter = Q(status__iexact='Success') | Q(status__iexact='Completed')
        if hasattr(Order, 'is_paid'):
            orders_filter |= Q(is_paid=True)

        orders_qs = Order.objects.filter(orders_filter)

        date_field = next((df for df in ['created_at', 'order_date', 'date', 'timestamp'] if hasattr(Order, df)), None)
        if date_field:
            orders_qs = apply_date_filter(orders_qs, date_field, date_filter, start_date, end_date).order_by(f'-{date_field}')

        amount_field = next((af for af in ['base_amount_inr', 'charged_amount', 'amount', 'total_amount', 'total', 'price'] if hasattr(Order, af)), None)
        total_revenue = orders_qs.aggregate(Sum(amount_field))[f'{amount_field}__sum'] or 0.0 if amount_field else 0.0
        mean_order = orders_qs.aggregate(Avg(amount_field))[f'{amount_field}__avg'] or 0.0 if amount_field else 0.0

        order_amounts = list(orders_qs.values_list(amount_field, flat=True)) if amount_field else []
        order_amounts = [float(a) for a in order_amounts if a is not None]
        median_order = calculate_median(order_amounts)

        headers = ['Sr. No.', 'Order ID', 'Customer', 'Amount Paid', 'Status', 'Order Date']
        if is_excel:
            headers += ['Payment Method', 'Transaction ID', 'Customer Email', 'Items Breakdown']

        details = []
        for idx, order in enumerate(orders_qs, start=1):
            order_id = f"#{getattr(order, 'id', '000')}"
            user_obj = getattr(order, 'user', getattr(order, 'student', getattr(order, 'customer', None)))
            user_str = user_obj.get_full_name() or user_obj.username if (user_obj and hasattr(user_obj, 'username')) else str(user_obj or 'N/A')
            
            val = get_order_amount(order)
            amount_str = f"₹{val:,.2f}"
            status_str = str(getattr(order, 'status', 'SUCCESS')).upper()
            raw_date = getattr(order, date_field, None) if date_field else None
            date_str = raw_date.strftime('%b %d, %Y') if raw_date else 'N/A'
            
            row = [idx, order_id, user_str, amount_str, status_str, date_str]

            if is_excel:
                payment_method = str(getattr(order, 'payment_method', getattr(order, 'payment_type', 'Razorpay / Card')))
                txn_id = str(getattr(order, 'payment_id', getattr(order, 'transaction_id', getattr(order, 'order_id', 'N/A'))))
                email_str = user_obj.email if (user_obj and hasattr(user_obj, 'email')) else 'N/A'
                
                items_list = []
                if hasattr(order, 'items'):
                    items_list = [getattr(item, 'book_title', getattr(getattr(item, 'book', None), 'title', 'Item')) for item in order.items.all()]
                elif hasattr(order, 'orderitem_set'):
                    items_list = [getattr(item, 'book_title', getattr(getattr(item, 'book', None), 'title', 'Item')) for item in order.orderitem_set.all()]
                elif hasattr(order, 'book') and order.book:
                    items_list = [getattr(order.book, 'title', 'Book Purchased')]
                
                items_str = ", ".join(items_list) if items_list else "Standard Order Items"

                row += [payment_method, txn_id, email_str, items_str]

            details.append({'row': row})

        summary_text = (
            f"Total Revenue: ₹{total_revenue:,.2f} | "
            f"Mean Order Value: ₹{mean_order:,.2f} | "
            f"Median Order Value: ₹{median_order:,.2f}"
        )

        return {
            'title': 'Orders Summary & Sales Analytics',
            'summary': summary_text,
            'headers': headers,
            'details': details
        }

    # 3. MOST PURCHASED ITEM ANALYTICS
    elif report_type in ['most_purchased_analytics', 'most_purchased', 'Most Purchased Items']:
        date_field = next((df for df in ['created_at', 'purchase_date', 'date'] if hasattr(PurchasedBook, df)), None)
        
        filter_kwargs = {}
        if date_field and date_filter != 'all':
            now = timezone.now()
            if date_filter == 'today':
                filter_kwargs[f'purchasedbook__{date_field}__date'] = now.date()
            elif date_filter == 'this_week':
                filter_kwargs[f'purchasedbook__{date_field}__gte'] = now - timezone.timedelta(days=7)
            elif date_filter == 'this_month':
                filter_kwargs[f'purchasedbook__{date_field}__month'] = now.month
                filter_kwargs[f'purchasedbook__{date_field}__year'] = now.year
            elif date_filter == 'this_year':
                filter_kwargs[f'purchasedbook__{date_field}__year'] = now.year

        count_filter = Q(**filter_kwargs) if filter_kwargs else None
        books_qs = Book.objects.annotate(
            total_sold=Count('purchasedbook', filter=count_filter)
        ).order_by('-total_sold', 'title')

        headers = ['Sr. No.', 'Book Number', 'Book Title', 'Author', 'Price', 'Units Sold', 'Demand Tag']
        if is_excel:
            headers += ['Estimated Sales Revenue (₹)', 'Inventory Availability']

        details = []
        for idx, book in enumerate(books_qs, start=1):
            book_no = getattr(book, 'isbn', getattr(book, 'id', 'N/A'))
            title = getattr(book, 'title', 'Unknown Title')
            author = getattr(book, 'author', 'N/A')
            p_val = float(getattr(book, 'price', 0.0) or 0.0)
            price = f"₹{p_val:,.2f}"
            count = book.total_sold
            demand_tag = 'HIGH DEMAND' if count >= 5 else ('MODERATE' if count > 0 else 'NO SALES')
            
            row = [idx, str(book_no), title, str(author), price, f"{count} Copies", demand_tag]

            if is_excel:
                est_revenue = f"₹{(count * p_val):,.2f}"
                stock_status = 'In Stock' if getattr(book, 'is_available', True) else 'Restock Required'
                row += [est_revenue, stock_status]

            details.append({'row': row})

        return {
            'title': 'Most Purchased Items',
            'summary': f'Total Catalog Items Tracked: {len(details)}',
            'headers': headers,
            'details': details
        }

    # 4. LEAST PURCHASED ITEM ANALYTICS
    elif report_type in ['least_purchased_analytics', 'least_purchased', 'Least Purchased Items']:
        date_field = next((df for df in ['created_at', 'purchase_date', 'date'] if hasattr(PurchasedBook, df)), None)
        
        filter_kwargs = {}
        if date_field and date_filter != 'all':
            now = timezone.now()
            if date_filter == 'today':
                filter_kwargs[f'purchasedbook__{date_field}__date'] = now.date()
            elif date_filter == 'this_week':
                filter_kwargs[f'purchasedbook__{date_field}__gte'] = now - timezone.timedelta(days=7)
            elif date_filter == 'this_month':
                filter_kwargs[f'purchasedbook__{date_field}__month'] = now.month
                filter_kwargs[f'purchasedbook__{date_field}__year'] = now.year
            elif date_filter == 'this_year':
                filter_kwargs[f'purchasedbook__{date_field}__year'] = now.year

        count_filter = Q(**filter_kwargs) if filter_kwargs else None
        books_qs = Book.objects.annotate(
            total_sold=Count('purchasedbook', filter=count_filter)
        ).order_by('total_sold', 'title')

        headers = ['Sr. No.', 'Book Number', 'Book Title', 'Author', 'Price', 'Units Sold', 'Demand Tag']
        if is_excel:
            headers += ['Estimated Sales Revenue (₹)', 'Inventory Availability']

        details = []
        for idx, book in enumerate(books_qs, start=1):
            book_no = getattr(book, 'isbn', getattr(book, 'id', 'N/A'))
            title = getattr(book, 'title', 'Unknown Title')
            author = getattr(book, 'author', 'N/A')
            p_val = float(getattr(book, 'price', 0.0) or 0.0)
            price = f"₹{p_val:,.2f}"
            count = book.total_sold
            demand_tag = 'LOW DEMAND' if count < 3 else 'MODERATE'
            
            row = [idx, str(book_no), title, str(author), price, f"{count} Copies", demand_tag]

            if is_excel:
                est_revenue = f"₹{(count * p_val):,.2f}"
                stock_status = 'In Stock' if getattr(book, 'is_available', True) else 'Out of Stock'
                row += [est_revenue, stock_status]

            details.append({'row': row})

        return {
            'title': 'Least Purchased Items',
            'summary': f'Low-Performing / Unsold Items: {len(details)}',
            'headers': headers,
            'details': details
        }

    # 5. MOST VIEWED ITEM
    elif report_type in ['most_viewed', 'most_viewed_analytics']:
        view_field = next((vf for vf in ['views', 'views_count', 'view_count', 'click_count'] if hasattr(Book, vf)), None)
        qs = Book.objects.all()

        if view_field:
            qs = qs.annotate(actual_views=Coalesce(F(view_field), 0)).order_by('-actual_views')
        else:
            qs = qs.annotate(actual_views=Coalesce(F('id') * 0, 0))

        headers = ['Sr. No.', 'Book Number', 'Book Title', 'Author', 'Price', 'Views', 'Popularity']
        if is_excel:
            headers += ['Category', 'Available Stock']

        details = []
        for idx, obj in enumerate(qs, start=1):
            book_no = getattr(obj, 'isbn', getattr(obj, 'id', 'N/A'))
            title = getattr(obj, 'title', 'Unknown Item')
            creator = getattr(obj, 'author', 'N/A')
            views_val = getattr(obj, 'actual_views', 0)

            price = float(getattr(obj, 'price', 0.0) or 0.0)
            price_str = f"₹{price:,.2f}"
            views_str = f"{views_val} Views"

            if views_val >= 50:
                status_str = 'POPULAR'
            elif views_val >= 10:
                status_str = 'MODERATE'
            else:
                status_str = 'LOW DEMAND'

            row = [idx, str(book_no), title, str(creator or 'N/A'), price_str, views_str, status_str]

            if is_excel:
                category = str(getattr(obj, 'category', 'General'))
                stock = 'In Stock' if getattr(obj, 'is_available', True) else 'Out of Stock'
                row += [category, stock]

            details.append({'row': row})

        return {
            'title': 'Most Viewed Items',
            'summary': f'Total Catalog Items: {len(details)}',
            'headers': headers,
            'details': details
        }

    # 6. LAST 5 ORDERS
    elif report_type == 'last_5_orders':
        date_field = next((df for df in ['created_at', 'order_date', 'date', 'timestamp'] if hasattr(Order, df)), None)
        orders_filter = Q(status__iexact='Success') | Q(status__iexact='Completed')
        if hasattr(Order, 'is_paid'):
            orders_filter |= Q(is_paid=True)

        orders_qs = Order.objects.filter(orders_filter)
        if date_field:
            orders_qs = orders_qs.order_by(f'-{date_field}')
        
        last_orders = orders_qs[:5]

        headers = ['Sr. No.', 'Order ID', 'Customer', 'Amount Paid', 'Status', 'Order Date']
        if is_excel:
            headers += ['Payment Method', 'Customer Email']

        details = []
        for idx, order in enumerate(last_orders, start=1):
            order_id = f"#{getattr(order, 'id', '000')}"
            user_obj = getattr(order, 'user', getattr(order, 'student', getattr(order, 'customer', None)))
            user_str = user_obj.get_full_name() or user_obj.username if (user_obj and hasattr(user_obj, 'username')) else str(user_obj or 'N/A')
            
            val = get_order_amount(order)
            amount_str = f"₹{val:,.2f}"
            status_str = str(getattr(order, 'status', 'SUCCESS')).upper()
            raw_date = getattr(order, date_field, None) if date_field else None
            date_str = raw_date.strftime('%b %d, %Y') if raw_date else 'N/A'
            
            row = [idx, order_id, user_str, amount_str, status_str, date_str]

            if is_excel:
                pm = str(getattr(order, 'payment_method', 'Razorpay / Online'))
                email = user_obj.email if (user_obj and hasattr(user_obj, 'email')) else 'N/A'
                row += [pm, email]

            details.append({'row': row})

        return {
            'title': 'Last 5 Orders',
            'summary': 'Recent Activity: 5 Orders',
            'headers': headers,
            'details': details
        }

    # 7. TOTAL AMOUNT PURCHASED
    elif report_type == 'total_amount_purchased':
        orders_filter = Q(status__iexact='Success') | Q(status__iexact='Completed')
        if hasattr(Order, 'is_paid'):
            orders_filter |= Q(is_paid=True)

        orders_qs = Order.objects.filter(orders_filter)
        date_field = next((df for df in ['created_at', 'order_date', 'date', 'timestamp'] if hasattr(Order, df)), None)
        if date_field:
            orders_qs = apply_date_filter(orders_qs, date_field, date_filter, start_date, end_date)

        amount_field = next((af for af in ['base_amount_inr', 'charged_amount', 'amount', 'total_amount', 'total', 'price'] if hasattr(Order, af)), None)
        total_revenue = orders_qs.aggregate(Sum(amount_field))[f'{amount_field}__sum'] or 0.0 if amount_field else 0.0
        mean_order = orders_qs.aggregate(Avg(amount_field))[f'{amount_field}__avg'] or 0.0 if amount_field else 0.0
        total_count = orders_qs.count()

        order_amounts = list(orders_qs.values_list(amount_field, flat=True)) if amount_field else []
        order_amounts = [float(a) for a in order_amounts if a is not None]
        median_order = calculate_median(order_amounts)

        headers = ['Sr. No.', 'Metric Name', 'Scope / Timeframe', 'Value / Amount', 'Status', 'Category']
        if is_excel:
            headers += ['Export Timestamp']

        now_str = timezone.now().strftime('%Y-%m-%d %H:%M')
        timeframe_label = date_filter.replace('_', ' ').title()

        details = [
            {'row': [1, 'Total Sales Volume', timeframe_label, f"{total_count} Orders", 'SUCCESS', 'Revenue Metric'] + ([now_str] if is_excel else [])},
            {'row': [2, 'Gross Sales Revenue', timeframe_label, f"₹{total_revenue:,.2f}", 'ACTIVE', 'Financial Metric'] + ([now_str] if is_excel else [])},
            {'row': [3, 'Mean Order Value (Average)', timeframe_label, f"₹{mean_order:,.2f}", 'ACTIVE', 'Statistical Metric'] + ([now_str] if is_excel else [])},
            {'row': [4, 'Median Order Value (Middle)', timeframe_label, f"₹{median_order:,.2f}", 'ACTIVE', 'Statistical Metric'] + ([now_str] if is_excel else [])}
        ]

        return {
            'title': 'Total Purchase & Statistical Revenue Analytics',
            'summary': f'Gross Revenue: ₹{total_revenue:,.2f} | Mean: ₹{mean_order:,.2f} | Median: ₹{median_order:,.2f}',
            'headers': headers,
            'details': details
        }

    # 8. PENDING BORROW REQUESTS
    elif report_type == 'pending_approvals':
        date_field = 'requested_at' if hasattr(BookRequest, 'requested_at') else ('created_at' if hasattr(BookRequest, 'created_at') else None)
        reqs_qs = BookRequest.objects.filter(status='PENDING')
        if date_field:
            reqs_qs = apply_date_filter(reqs_qs, date_field, date_filter, start_date, end_date).order_by(f'-{date_field}')

        headers = ['Sr. No.', 'Book Number', 'Book Title', 'Author', 'Price', 'Requested By', 'Status', 'Requested Date']
        if is_excel:
            headers += ['User Email', 'User Phone / Contact', 'Max Allowed Borrow Days']

        details = []
        for idx, req in enumerate(reqs_qs, start=1):
            book_obj = getattr(req, 'book', None)
            book_no = getattr(book_obj, 'isbn', getattr(book_obj, 'id', 'N/A')) if book_obj else 'N/A'
            book_title = getattr(book_obj, 'title', 'Unknown Book') if book_obj else 'Unknown Book'
            author_name = getattr(book_obj, 'author', getattr(book_obj, 'author_name', 'N/A')) if book_obj else 'N/A'
            price = f"₹{float(getattr(book_obj, 'price', 0.0) or 0.0):,.2f}" if book_obj and hasattr(book_obj, 'price') else 'N/A'
            
            user_obj = getattr(req, 'student', getattr(req, 'user', None))
            user_display = getattr(user_obj, 'username', str(user_obj or 'N/A'))
            
            raw_date = getattr(req, date_field, None) if date_field else None
            requested_date_str = raw_date.strftime('%b %d, %Y') if raw_date else 'N/A'
            
            row = [idx, str(book_no), book_title, author_name, price, user_display, 'PENDING', requested_date_str]

            if is_excel:
                user_email = getattr(user_obj, 'email', 'N/A')
                user_phone = getattr(user_obj, 'phone_number', getattr(getattr(user_obj, 'profile', None), 'phone', 'N/A'))
                max_days = getattr(req, 'max_days', 14)
                row += [user_email, str(user_phone), f"{max_days} Days"]

            details.append({'row': row})

        return {
            'title': 'Pending Borrow Requests',
            'summary': f'Total Pending Requests: {len(details)}',
            'headers': headers,
            'details': details
        }

    # 9. APPROVED BORROW REQUESTS
    elif report_type == 'approved_books':
        date_field = 'requested_at' if hasattr(BookRequest, 'requested_at') else ('created_at' if hasattr(BookRequest, 'created_at') else None)
        reqs_qs = BookRequest.objects.filter(status='APPROVED')
        if date_field:
            reqs_qs = apply_date_filter(reqs_qs, date_field, date_filter, start_date, end_date).order_by(f'-{date_field}')

        headers = ['Sr. No.', 'Book Number', 'Book Title', 'Author', 'Price', 'Approved To', 'Status', 'Approval Date']
        if is_excel:
            headers += ['Approved By (Librarian)', 'Due Date', 'Return Status']

        details = []
        for idx, req in enumerate(reqs_qs, start=1):
            book_obj = getattr(req, 'book', None)
            book_no = getattr(book_obj, 'isbn', getattr(book_obj, 'id', 'N/A')) if book_obj else 'N/A'
            book_title = getattr(book_obj, 'title', 'Unknown Book') if book_obj else 'Unknown Book'
            author_name = getattr(book_obj, 'author', getattr(book_obj, 'author_name', 'N/A')) if book_obj else 'N/A'
            price = f"₹{float(getattr(book_obj, 'price', 0.0) or 0.0):,.2f}" if book_obj and hasattr(book_obj, 'price') else 'N/A'
            
            user_obj = getattr(req, 'student', getattr(req, 'user', None))
            user_display = user_obj.get_full_name() or user_obj.username if (user_obj and hasattr(user_obj, 'username')) else str(user_obj or 'N/A')
            
            raw_date = getattr(req, date_field, None) if date_field else None
            approval_date_str = raw_date.strftime('%b %d, %Y') if raw_date else 'N/A'
            
            row = [idx, str(book_no), book_title, author_name, price, user_display, 'APPROVED', approval_date_str]

            if is_excel:
                approved_by = getattr(req, 'approved_by', 'System / Admin')
                due_date_raw = getattr(req, 'due_date', None)
                due_date_str = due_date_raw.strftime('%b %d, %Y') if due_date_raw else 'N/A'
                is_returned = 'Returned' if getattr(req, 'is_returned', False) else 'Currently Borrowed'

                row += [str(approved_by), due_date_str, is_returned]

            details.append({'row': row})

        return {
            'title': 'Approved Borrow Requests',
            'summary': f'Total Approved Requests: {len(details)}',
            'headers': headers,
            'details': details
        }

    return None


@login_required
@user_passes_test(is_authorized_staff)
def reports_dashboard(request):
    total_users = User.objects.count()

    amount_field = next((af for af in ['base_amount_inr', 'charged_amount', 'amount', 'total_amount', 'total', 'price'] if hasattr(Order, af)), None)
    
    orders_filter = Q(status__iexact='Success') | Q(status__iexact='Completed')
    if hasattr(Order, 'is_paid'):
        orders_filter |= Q(is_paid=True)

    successful_orders = Order.objects.filter(orders_filter)
    
    if amount_field:
        agg_data = successful_orders.aggregate(total_sum=Sum(amount_field), avg_val=Avg(amount_field))
        revenue_sum = agg_data['total_sum'] or 0.0
        avg_order_val = agg_data['avg_val'] or 0.0
    else:
        revenue_sum = 0.0
        avg_order_val = 0.0

    most_purchased_item = "N/A"
    most_purchased_qs = PurchasedBook.objects.values('book__title').annotate(total=Count('id')).order_by('-total').first()
    if most_purchased_qs and most_purchased_qs.get('book__title'):
        most_purchased_item = most_purchased_qs['book__title']

    least_purchased_item = "N/A"
    least_purchased_qs = PurchasedBook.objects.values('book__title').annotate(total=Count('id')).order_by('total').first()
    if least_purchased_qs and least_purchased_qs.get('book__title'):
        least_purchased_item = least_purchased_qs['book__title']

    most_viewed_item = "N/A"
    view_field = next((vf for vf in ['views', 'views_count', 'view_count', 'click_count'] if hasattr(Book, vf)), None)
    if view_field:
        top_viewed_book = Book.objects.order_by(f'-{view_field}').first()
        if top_viewed_book and getattr(top_viewed_book, 'title', None):
            most_viewed_item = top_viewed_book.title

    recent_orders_count = successful_orders.count()

    order_date_field = next((df for df in ['created_at', 'order_date', 'date', 'timestamp'] if hasattr(Order, df)), None)
    
    month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    monthly_data = [0.0] * 12
    monthly_median_data = [0.0] * 12

    quarter_labels = ['Q1', 'Q2', 'Q3', 'Q4']
    quarterly_data = [0.0] * 4
    quarterly_median_data = [0.0] * 4

    yearly_data = []
    yearly_median_data = []
    year_labels = []

    if amount_field and order_date_field:
        current_year = timezone.now().year

        # --- Month-wise Aggregation ---
        monthly_qs = successful_orders.filter(**{f"{order_date_field}__year": current_year}).annotate(
            period=ExtractMonth(order_date_field)
        ).values('period').annotate(total=Sum(amount_field)).order_by('period')

        for item in monthly_qs:
            if item['period']:
                monthly_data[item['period'] - 1] = float(item['total'] or 0.0)

        for m_idx in range(1, 13):
            m_orders = successful_orders.filter(
                **{f"{order_date_field}__year": current_year, f"{order_date_field}__month": m_idx}
            )
            m_vals = [float(v) for v in m_orders.values_list(amount_field, flat=True) if v is not None]
            monthly_median_data[m_idx - 1] = calculate_median(m_vals)

        # --- Quarter-wise Aggregation ---
        quarterly_qs = successful_orders.annotate(
            period=ExtractQuarter(order_date_field)
        ).values('period').annotate(total=Sum(amount_field)).order_by('period')

        for item in quarterly_qs:
            if item['period']:
                quarterly_data[item['period'] - 1] = float(item['total'] or 0.0)

        for q_idx in range(1, 5):
            q_orders = successful_orders.filter(**{f"{order_date_field}__quarter": q_idx})
            q_vals = [float(v) for v in q_orders.values_list(amount_field, flat=True) if v is not None]
            quarterly_median_data[q_idx - 1] = calculate_median(q_vals)

        # --- Year-wise Aggregation ---
        yearly_qs = successful_orders.annotate(
            period=ExtractYear(order_date_field)
        ).values('period').annotate(total=Sum(amount_field)).order_by('period')

        for item in yearly_qs:
            if item['period']:
                y_val = item['period']
                year_labels.append(str(y_val))
                yearly_data.append(float(item['total'] or 0.0))

                y_orders = successful_orders.filter(**{f"{order_date_field}__year": y_val})
                y_vals = [float(v) for v in y_orders.values_list(amount_field, flat=True) if v is not None]
                yearly_median_data.append(calculate_median(y_vals))

    if not year_labels:
        year_labels = ['2026']
        yearly_data = [0.0]
        yearly_median_data = [0.0]

    revenue_analytics = {
        'month': {
            'labels': month_labels,
            'revenue': monthly_data,
            'median': monthly_median_data,
            'data': monthly_data
        },
        'quarter': {
            'labels': quarter_labels,
            'revenue': quarterly_data,
            'median': quarterly_median_data,
            'data': quarterly_data
        },
        'year': {
            'labels': year_labels,
            'revenue': yearly_data,
            'median': yearly_median_data,
            'data': yearly_data
        },
    }

    context = {
        'total_users': total_users,
        'total_revenue': f"{revenue_sum:,.2f}",
        'avg_order_value': f"{avg_order_val:,.2f}",
        'most_purchased_item': most_purchased_item,
        'least_purchased_item': least_purchased_item,
        'most_viewed_item': most_viewed_item,
        'recent_orders_count': min(recent_orders_count, 5),
        'revenue_analytics': revenue_analytics,
    }
    return render(request, 'reports/dashboard.html', context)


@login_required
@user_passes_test(is_authorized_staff)
def get_report_data(request):
    report_type = request.GET.get('type') or request.GET.get('report_type', 'most_viewed')
    date_filter = request.GET.get('filter') or request.GET.get('date_range', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    page = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 10)

    try:
        page_size = max(1, int(page_size))
    except ValueError:
        page_size = 10

    data = fetch_report_payload(report_type, date_filter, start_date, end_date, is_excel=False)
    if data:
        all_details = data.get('details', [])
        paginator = Paginator(all_details, page_size)

        try:
            paginated_page = paginator.page(page)
        except PageNotAnInteger:
            paginated_page = paginator.page(1)
        except EmptyPage:
            paginated_page = paginator.page(paginator.num_pages)

        pagination_info = {
            'current_page': paginated_page.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_next': paginated_page.has_next(),
            'has_previous': paginated_page.has_previous(),
            'page_size': page_size
        }

        return JsonResponse({
            'status': 'success',
            'title': data['title'],
            'summary': data['summary'],
            'headers': data['headers'],
            'details': list(paginated_page.object_list),
            'pagination': pagination_info
        })

    return JsonResponse({'status': 'error', 'message': 'Invalid report type selected.'}, status=400)


@login_required
@user_passes_test(is_authorized_staff)
def export_report_excel(request):
    report_type = request.GET.get('type') or request.GET.get('report_type')
    date_filter = request.GET.get('filter') or request.GET.get('date_range', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    data = fetch_report_payload(report_type, date_filter, start_date, end_date, is_excel=True)
    if not data:
        return HttpResponse("Invalid Report Type Selected", status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "System Report"

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="1E3A8A")
    summary_font = Font(name="Arial", size=10, italic=True, color="475569")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    filter_label = f"{start_date} to {end_date}" if date_filter == 'custom' and start_date and end_date else date_filter.replace('_', ' ').title()

    ws.append([data['title']])
    ws.append([f"Summary: {data['summary']}"])
    ws.append([f"Date Filter: {filter_label}"])
    ws.append([])

    ws['A1'].font = title_font
    ws['A2'].font = summary_font
    ws['A3'].font = summary_font

    ws.append(data['headers'])
    header_row_idx = 5
    for cell in ws[header_row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for detail in data['details']:
        ws.append(detail['row'])

    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, min_col=1, max_col=len(data['headers'])):
        for cell in row:
            cell.border = border_thin
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{date_filter}.xlsx"'
    wb.save(response)

    return response